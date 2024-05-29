import os
import shutil
from time import sleep

import pandas as pd
from PyPDF2 import PdfMerger
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Side, Border, Font
from win32com import client

from DataBase.INDXTBDB.AmplAccIndxTableDB import AmplAccIndxTableDB
from DataBase.INDXTBDB.DOAMeasIndxTableDB import DoaMeasIndxTableDB
from DataBase.INDXTBDB.FreqAccIndxTableDB import FreqAccIndxTableDB
from DataBase.INDXTBDB.PRIMeasIndxTableDB import PriMeasIndxTableDB
from DataBase.INDXTBDB.PWMeasIndxTableDB import PwMeasIndxTableDB
from DataBase.INDXTBDB.SensMeasAccIndxTableDB import SensMeasIndxTableDB
from DataBase.TESTTBDB.AmplAccTestTableDB import AmplAccTestTableDB
from DataBase.TESTTBDB.DoaMeasTestTableDB import DoaMeasTestTableDB
from DataBase.TESTTBDB.FreqAccTestTableDB import FreqAccTestTableDB
from DataBase.TESTTBDB.PriMeasTestTableDB import PriMeasTestTableDB
from DataBase.TESTTBDB.PwMeasTestTableDB import PwMeasTestTableDB
from DataBase.TESTTBDB.SensMeasTestTableDB import SensMeasTestTableDB
from Reports.ExcelCells.FrequencyTDR import *


class ReportPdf():
    def __init__(self):
        self.ReportPath = "C:/Users/jaswa/PycharmProjects/ATEC1/Reports/"
        self.MainsrcPath = "C:/Users/jaswa/PycharmProjects/ATEC1/MainSrc/"
    def PdfPreview(self, TableSelected=['esmfreqacctest_2024_04_30_15_50_45']):
        #######################################Creating Excel Files#####################################################
        self.TableSelection = TableSelected
        key_matched = []
        testnames = ["freqacc", "pwacc", "primeas", "doameas", "amplacc", "sensmeas"]
        ForTablename = self.TableSelection
        for key in testnames:
            for file in ForTablename:
                if key in file:
                    key_matched.append(key)
                    print(key)
        print(key_matched)
        for i in range(0, len(key_matched)):
            print(key_matched[i])
            if key_matched[i] == 'freqacc':
                excelfilename = self.TableSelection[i]
                GetFreqIndxTable = FreqAccIndxTableDB()
                GetFreqIndxTable.GetFreqRowRecord(select_record=self.TableSelection[i])
                dataappend = GetFreqIndxTable.GetFreqIndxRowRecord
                print(dataappend)
                path = f'{self.ReportPath}Templates/FreqAccTestTemplate.xlsx'
                GetFreqTestTable = FreqAccTestTableDB()
                GetFreqTestTable.GetFreqAccTestTable(testtablename=self.TableSelection[i])
                dfreadtoui = GetFreqTestTable.CurDbTableFreq
                # print(dfreadtoui)
                self.settestvalues = dfreadtoui["set_freq"]
                self.meastestvalues = dfreadtoui["meas_freq"]
                self.errortestvalues = dfreadtoui["error"]
                dfRowFil = dataappend
                workbook = load_workbook(path)
                sheetonename = "sheet1"
                sheet1 = workbook.active
                sheet1.title = sheetonename
                sheet1[START_FREQUENCY] = dfRowFil.iloc[0][6]  # Start Freq
                sheet1[STOP_FREQUENCY] = (dfRowFil.iloc[0][7])  # Stop Freq
                sheet1[STEP_FREQUENCY] = (dfRowFil.iloc[0][8])  # Step Freq
                sheet1[FREQ_POS_ANGLE] = (dfRowFil.iloc[0][11])  # Position Angle
                sheet1[FREQ_PW] = (dfRowFil.iloc[0][13])  # PW
                sheet1[FREQ_PRI] = (dfRowFil.iloc[0][14])  # PRI
                sheet1[FREQ_SET_POWER] = (dfRowFil.iloc[0][10])  # Set Power
                datetrim = dataappend.iloc[0][0]
                dateformat = str(datetrim)[:10]
                timeformat = str(datetrim)[11:19]
                sheet1[TEST_SYSTEM] = (dataappend.iloc[0][3]).upper()  # system
                sheet1[TEST_MODE] = dataappend.iloc[0][4]  # mode
                sheet1[USER_NAME] = dataappend.iloc[0][1]  # name
                sheet1[TEST_DATE] = dateformat  # date
                sheet1[TEST_TIME] = timeformat  # time
                sheet1[SIGNAL_CATEGORY] = (dataappend.iloc[0][12])  # Signal Category
                sheet1[SYSTEM_NUMBER] = dataappend.iloc[0][2]  # System number

                for i in range(0, len(dfreadtoui)):
                    sheet1.merge_cells(start_row=18 + i, start_column=3, end_row=18 + i,
                                       end_column=4)  # Merging C and D cols
                    sheet1.merge_cells(start_row=18 + i, start_column=5, end_row=18 + i,
                                       end_column=7)  # Merging E to G cols
                    sheet1.merge_cells(start_row=18 + i, start_column=8, end_row=18 + i,
                                       end_column=9)  # Merging H and I cols
                    self.set_border(sheet1, f'''B{18 + i}:I{18 + i}''')  # for excel sheet table cell borders

                    sheet1[f'''B{18 + i}'''] = i + 1
                    sheet1[f'''B{18 + i}'''].alignment = Alignment(
                        horizontal='center')  # Alignment of text center, left or right
                    sheet1[f'''C{18 + i}'''] = self.settestvalues.values[i]
                    sheet1[f'''C{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''E{18 + i}'''] = self.meastestvalues.values[i]
                    sheet1[f'''E{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''H{18 + i}'''] = self.errortestvalues.values[i]
                    sheet1[f'''H{18 + i}'''].alignment = Alignment(horizontal='center')

                    # merging cells for text at the end of sheet
                self.tableposition = 18 + len(dfreadtoui) + 2
                sheet1.merge_cells(start_row=self.tableposition, start_column=2, end_row=self.tableposition + 1,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 3, start_column=2, end_row=self.tableposition + 4,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 6, start_column=2, end_row=self.tableposition + 7,
                                   end_column=5)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=2, end_row=self.tableposition + 9,
                                   end_column=3)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=8, end_row=self.tableposition + 9,
                                   end_column=10)

                # excel writing
                sheet1[f'B{self.tableposition}'] = 'RMS Error :'
                sheet1[f'B{self.tableposition + 3}'] = 'Specification :'
                sheet1[f'B{self.tableposition + 6}'] = 'Status :'
                sheet1[f'B{self.tableposition + 8}'] = 'Tested By'
                sheet1[f'H{self.tableposition + 8}'] = 'Verified By'
                self.FontSizeStyle = [f'B{self.tableposition}', f'B{self.tableposition + 3}',
                                      f'B{self.tableposition + 6}',
                                      f'B{self.tableposition + 8}', f'H{self.tableposition + 8}']
                for i in range(0, len(self.FontSizeStyle)):
                    sheet1[self.FontSizeStyle[i]].font = Font(size=12, bold=True)  # font size and style
                # save the file
                outfile = excelfilename + '.xlsx'
                workbook.save(filename=f'{self.ReportPath}temp/' + outfile)
            elif key_matched[i] == 'pwacc':
                excelfilename = self.TableSelection[i]
                GetPwIndxTable = PwMeasIndxTableDB()
                GetPwIndxTable.GetPwRowRecord(select_record=self.TableSelection[i])
                dataappend = GetPwIndxTable.GetPwIndxRowRecord
                path = f'{self.ReportPath}Templates/PwAccTestTemplate.xlsx'
                GetPwTestTable = PwMeasTestTableDB()
                GetPwTestTable.GetPwMeasTestTable(testtablename=self.TableSelection[i])
                dfreadtoui = GetPwTestTable.CurDbTablePw
                self.settestvalues = dfreadtoui["set_pw"]
                self.meastestvalues = dfreadtoui["meas_pw"]
                self.errortestvalues = dfreadtoui["error"]
                dfRowFil = dataappend
                workbook = load_workbook(path)
                sheetonename = "sheet1"
                sheet1 = workbook.active
                sheet1.title = sheetonename
                sheet1[START_PW] = dfRowFil.iloc[0][6]  # Start PW
                sheet1[STOP_PW] = (dfRowFil.iloc[0][7])  # Stop PW
                sheet1[STEP_PW] = (dfRowFil.iloc[0][8])  # Step PW
                sheet1[PW_POS_ANGLE] = (dfRowFil.iloc[0][11])  # Position Angle
                sheet1[PW_FREQ] = (dfRowFil.iloc[0][13])  # FREQ
                sheet1[PW_PRI] = (dfRowFil.iloc[0][14])  # PRI
                sheet1[PW_SET_POWER] = (dfRowFil.iloc[0][10])  # Set Power
                datetrim = dataappend.iloc[0][0]
                dateformat = str(datetrim)[:10]
                timeformat = str(datetrim)[11:19]
                sheet1[TEST_SYSTEM] = (dataappend.iloc[0][3]).upper()  # system
                sheet1[TEST_MODE] = dataappend.iloc[0][4]  # mode
                sheet1[USER_NAME] = dataappend.iloc[0][1]  # name
                sheet1[TEST_DATE] = dateformat  # date
                sheet1[TEST_TIME] = timeformat  # time
                sheet1[SIGNAL_CATEGORY] = (dataappend.iloc[0][12])  # Signal Category
                sheet1[SYSTEM_NUMBER] = dataappend.iloc[0][2]  # System number
                ############
                for i in range(0, len(dfreadtoui)):
                    sheet1.merge_cells(start_row=18 + i, start_column=3, end_row=18 + i,
                                       end_column=4)  # Merging C and D cols
                    sheet1.merge_cells(start_row=18 + i, start_column=5, end_row=18 + i,
                                       end_column=7)  # Merging E to G cols
                    sheet1.merge_cells(start_row=18 + i, start_column=8, end_row=18 + i,
                                       end_column=9)  # Merging H and I cols

                    self.set_border(sheet1, f'''B{18 + i}:I{18 + i}''')  # for excel sheet table cell borders

                    sheet1[f'''B{18 + i}'''] = i + 1
                    sheet1[f'''B{18 + i}'''].alignment = Alignment(
                        horizontal='center')  # Alignment of text center, left or right
                    sheet1[f'''C{18 + i}'''] = self.settestvalues.values[i]
                    sheet1[f'''C{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''E{18 + i}'''] = self.meastestvalues.values[i]
                    sheet1[f'''E{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''H{18 + i}'''] = self.errortestvalues.values[i]
                    sheet1[f'''H{18 + i}'''].alignment = Alignment(horizontal='center')

                    # merging cells for text at the end of sheet
                self.tableposition = 18 + len(dfreadtoui) + 2
                sheet1.merge_cells(start_row=self.tableposition, start_column=2, end_row=self.tableposition + 1,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 3, start_column=2, end_row=self.tableposition + 4,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 6, start_column=2, end_row=self.tableposition + 7,
                                   end_column=5)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=2, end_row=self.tableposition + 9,
                                   end_column=3)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=8, end_row=self.tableposition + 9,
                                   end_column=10)

                # excel writing
                sheet1[f'B{self.tableposition}'] = 'RMS Error :'
                sheet1[f'B{self.tableposition + 3}'] = 'Specification :'
                sheet1[f'B{self.tableposition + 6}'] = 'Status :'
                sheet1[f'B{self.tableposition + 8}'] = 'Tested By'
                sheet1[f'H{self.tableposition + 8}'] = 'Verified By'
                self.FontSizeStyle = [f'B{self.tableposition}', f'B{self.tableposition + 3}',
                                      f'B{self.tableposition + 6}',
                                      f'B{self.tableposition + 8}', f'H{self.tableposition + 8}']
                for i in range(0, len(self.FontSizeStyle)):
                    sheet1[self.FontSizeStyle[i]].font = Font(size=12, bold=True)  # font size and style
                # save the file
                outfile = excelfilename + '.xlsx'
                workbook.save(filename=f'{self.ReportPath}temp/' + outfile)
            elif key_matched[i] == 'primeas':
                excelfilename = self.TableSelection[i]
                GetPriIndxTable = PriMeasIndxTableDB()
                GetPriIndxTable.GetPriRowRecord(select_record=self.TableSelection[i])
                dataappend = GetPriIndxTable.GetPriIndxRowRecord
                path = f'{self.ReportPath}Templates/PriAccTestTemplate.xlsx'
                GetPriTestTable = PriMeasTestTableDB()
                GetPriTestTable.GetPriMeasTestTable(testtablename=self.TableSelection[i])
                dfreadtoui = GetPriTestTable.CurDbTablePri
                self.settestvalues = dfreadtoui["set_pri"]
                self.meastestvalues = dfreadtoui["meas_pri"]
                self.errortestvalues = dfreadtoui["error"]
                dfRowFil = dataappend
                workbook = load_workbook(path)
                sheetonename = "sheet1"
                sheet1 = workbook.active
                sheet1.title = sheetonename
                sheet1[START_PRI] = dfRowFil.iloc[0][6]  # Start PW
                sheet1[STOP_PRI] = (dfRowFil.iloc[0][7])  # Stop PW
                sheet1[STEP_PRI] = (dfRowFil.iloc[0][8])  # Step PW
                sheet1[PRI_POS_ANGLE] = (dfRowFil.iloc[0][11])  # Position Angle
                sheet1[PRI_FREQ] = (dfRowFil.iloc[0][13])  # FREQ
                sheet1[PRI_PW] = (dfRowFil.iloc[0][14])  # PRI
                sheet1[PRI_SET_POWER] = (dfRowFil.iloc[0][10])  # Set Power
                datetrim = dataappend.iloc[0][0]
                dateformat = str(datetrim)[:10]
                timeformat = str(datetrim)[11:19]
                sheet1[TEST_SYSTEM] = (dataappend.iloc[0][3]).upper()  # system
                sheet1[TEST_MODE] = dataappend.iloc[0][4]  # mode
                sheet1[USER_NAME] = dataappend.iloc[0][1]  # name
                sheet1[TEST_DATE] = dateformat  # date
                sheet1[TEST_TIME] = timeformat  # time
                sheet1[SIGNAL_CATEGORY] = (dataappend.iloc[0][12])  # Signal Category
                sheet1[SYSTEM_NUMBER] = dataappend.iloc[0][2]  # System number
                ##########################################################
                for i in range(0, len(dfreadtoui)):
                    sheet1.merge_cells(start_row=18 + i, start_column=3, end_row=18 + i,
                                       end_column=4)  # Merging C and D cols
                    sheet1.merge_cells(start_row=18 + i, start_column=5, end_row=18 + i,
                                       end_column=7)  # Merging E to G cols
                    sheet1.merge_cells(start_row=18 + i, start_column=8, end_row=18 + i,
                                       end_column=9)  # Merging H and I cols

                    self.set_border(sheet1, f'''B{18 + i}:I{18 + i}''')  # for excel sheet table cell borders

                    sheet1[f'''B{18 + i}'''] = i + 1
                    sheet1[f'''B{18 + i}'''].alignment = Alignment(
                        horizontal='center')  # Alignment of text center, left or right
                    sheet1[f'''C{18 + i}'''] = self.settestvalues.values[i]
                    sheet1[f'''C{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''E{18 + i}'''] = self.meastestvalues.values[i]
                    sheet1[f'''E{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''H{18 + i}'''] = self.errortestvalues.values[i]
                    sheet1[f'''H{18 + i}'''].alignment = Alignment(horizontal='center')

                    # merging cells for text at the end of sheet
                self.tableposition = 18 + len(dfreadtoui) + 2
                sheet1.merge_cells(start_row=self.tableposition, start_column=2, end_row=self.tableposition + 1,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 3, start_column=2, end_row=self.tableposition + 4,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 6, start_column=2, end_row=self.tableposition + 7,
                                   end_column=5)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=2, end_row=self.tableposition + 9,
                                   end_column=3)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=8, end_row=self.tableposition + 9,
                                   end_column=10)

                # excel writing
                sheet1[f'B{self.tableposition}'] = 'RMS Error :'
                sheet1[f'B{self.tableposition + 3}'] = 'Specification :'
                sheet1[f'B{self.tableposition + 6}'] = 'Status :'
                sheet1[f'B{self.tableposition + 8}'] = 'Tested By'
                sheet1[f'H{self.tableposition + 8}'] = 'Verified By'
                self.FontSizeStyle = [f'B{self.tableposition}', f'B{self.tableposition + 3}',
                                      f'B{self.tableposition + 6}',
                                      f'B{self.tableposition + 8}', f'H{self.tableposition + 8}']
                for i in range(0, len(self.FontSizeStyle)):
                    sheet1[self.FontSizeStyle[i]].font = Font(size=12, bold=True)  # font size and style
                # save the file
                outfile = excelfilename + '.xlsx'
                workbook.save(filename=f'{self.ReportPath}temp/' + outfile)
                ########################################################################################################
            elif key_matched[i] == 'amplacc':
                excelfilename = self.TableSelection[i]
                GetAmplIndxTable = AmplAccIndxTableDB()
                GetAmplIndxTable.GetAmplRowRecord(select_record=self.TableSelection[i])
                dataappend = GetAmplIndxTable.GetAmplIndxRowRecord
                path = f'{self.ReportPath}Templates/AmplAccTestTemplate.xlsx'
                GetAmplTestTable = AmplAccTestTableDB()
                GetAmplTestTable.GetAmplAccTestTable(testtablename=self.TableSelection[i])
                dfreadtoui = GetAmplTestTable.CurDbTableAmpl
                self.settestvalues = dfreadtoui["set_ampl"]
                self.meastestvalues = dfreadtoui["meas_ampl"]
                self.errortestvalues = dfreadtoui["error"]
                dfRowFil = dataappend
                workbook = load_workbook(path)
                sheetonename = "sheet1"
                sheet1 = workbook.active
                sheet1.title = sheetonename
                sheet1[START_AMPL] = dfRowFil.iloc[0][6]  # Start PW
                sheet1[STOP_AMPL] = (dfRowFil.iloc[0][7])  # Stop PW
                sheet1[STEP_AMPL] = (dfRowFil.iloc[0][8])  # Step PW
                sheet1[AMPL_POS_ANGLE] = (dfRowFil.iloc[0][11])  # Position Angle
                sheet1[AMPL_PW] = (dfRowFil.iloc[0][13])  # FREQ
                sheet1[AMPL_PRI] = (dfRowFil.iloc[0][14])  # PRI
                sheet1[AMPL_FREQ] = (dfRowFil.iloc[0][10])  # Set Power
                datetrim = dataappend.iloc[0][0]
                dateformat = str(datetrim)[:10]
                timeformat = str(datetrim)[11:19]
                sheet1[TEST_SYSTEM] = (dataappend.iloc[0][3]).upper()  # system
                sheet1[TEST_MODE] = dataappend.iloc[0][4]  # mode
                sheet1[USER_NAME] = dataappend.iloc[0][1]  # name
                sheet1[TEST_DATE] = dateformat  # date
                sheet1[TEST_TIME] = timeformat  # time
                sheet1[SIGNAL_CATEGORY] = (dataappend.iloc[0][12])  # Signal Category
                sheet1[SYSTEM_NUMBER] = dataappend.iloc[0][2]  # System number
                ##########################################################
                for i in range(0, len(dfreadtoui)):
                    sheet1.merge_cells(start_row=18 + i, start_column=3, end_row=18 + i,
                                       end_column=4)  # Merging C and D cols
                    sheet1.merge_cells(start_row=18 + i, start_column=5, end_row=18 + i,
                                       end_column=7)  # Merging E to G cols
                    sheet1.merge_cells(start_row=18 + i, start_column=8, end_row=18 + i,
                                       end_column=9)  # Merging H and I cols

                    self.set_border(sheet1, f'''B{18 + i}:I{18 + i}''')  # for excel sheet table cell borders

                    sheet1[f'''B{18 + i}'''] = i + 1
                    sheet1[f'''B{18 + i}'''].alignment = Alignment(
                        horizontal='center')  # Alignment of text center, left or right
                    sheet1[f'''C{18 + i}'''] = self.settestvalues.values[i]
                    sheet1[f'''C{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''E{18 + i}'''] = self.meastestvalues.values[i]
                    sheet1[f'''E{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''H{18 + i}'''] = self.errortestvalues.values[i]
                    sheet1[f'''H{18 + i}'''].alignment = Alignment(horizontal='center')

                    # merging cells for text at the end of sheet
                self.tableposition = 18 + len(dfreadtoui) + 2
                sheet1.merge_cells(start_row=self.tableposition, start_column=2, end_row=self.tableposition + 1,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 3, start_column=2, end_row=self.tableposition + 4,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 6, start_column=2, end_row=self.tableposition + 7,
                                   end_column=5)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=2, end_row=self.tableposition + 9,
                                   end_column=3)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=8, end_row=self.tableposition + 9,
                                   end_column=10)

                # excel writing
                sheet1[f'B{self.tableposition}'] = 'RMS Error :'
                sheet1[f'B{self.tableposition + 3}'] = 'Specification :'
                sheet1[f'B{self.tableposition + 6}'] = 'Status :'
                sheet1[f'B{self.tableposition + 8}'] = 'Tested By'
                sheet1[f'H{self.tableposition + 8}'] = 'Verified By'
                self.FontSizeStyle = [f'B{self.tableposition}', f'B{self.tableposition + 3}',
                                      f'B{self.tableposition + 6}',
                                      f'B{self.tableposition + 8}', f'H{self.tableposition + 8}']
                for i in range(0, len(self.FontSizeStyle)):
                    sheet1[self.FontSizeStyle[i]].font = Font(size=12, bold=True)  # font size and style
                # save the file
                outfile = excelfilename + '.xlsx'
                workbook.save(filename=f'{self.ReportPath}temp/' + outfile)
                print(dataappend)
            elif key_matched[i] == 'doameas':
                excelfilename = self.TableSelection[i]
                GetDoaIndxTable = DoaMeasIndxTableDB()
                GetDoaIndxTable.GetDoaRowRecord(select_record=self.TableSelection[i])
                dataappend = GetDoaIndxTable.GetDoaIndxRowRecord
                path = f'{self.ReportPath}Templates/DoaAccTestTemplate.xlsx'
                GetDoaTestTable = DoaMeasTestTableDB()
                GetDoaTestTable.GetDoaMeasTestTable(testtablename=self.TableSelection[i])
                dfreadtoui = GetDoaTestTable.CurDbTableDoa
                self.settestvalues = dfreadtoui["set_angle"]
                self.meastestvalues = dfreadtoui["meas_angle"]
                self.errortestvalues = dfreadtoui["error"]
                dfRowFil = dataappend
                workbook = load_workbook(path)
                sheetonename = "sheet1"
                sheet1 = workbook.active
                sheet1.title = sheetonename
                sheet1[START_DOA] = dfRowFil.iloc[0][6]  # Start PW
                sheet1[STOP_DOA] = (dfRowFil.iloc[0][7])  # Stop PW
                sheet1[STEP_DOA] = (dfRowFil.iloc[0][8])  # Step PW
                sheet1[DOA_PW] = (dfRowFil.iloc[0][11])  # Position Angle
                sheet1[DOA_FREQ] = (dfRowFil.iloc[0][13])  # FREQ
                sheet1[DOA_PRI] = (dfRowFil.iloc[0][14])  # PRI
                sheet1[DOA_SET_POWER] = (dfRowFil.iloc[0][10])  # Set Power
                datetrim = dataappend.iloc[0][0]
                dateformat = str(datetrim)[:10]
                timeformat = str(datetrim)[11:19]
                sheet1[TEST_SYSTEM] = (dataappend.iloc[0][3]).upper()  # system
                sheet1[TEST_MODE] = dataappend.iloc[0][4]  # mode
                sheet1[USER_NAME] = dataappend.iloc[0][1]  # name
                sheet1[TEST_DATE] = dateformat  # date
                sheet1[TEST_TIME] = timeformat  # time
                sheet1[SIGNAL_CATEGORY] = (dataappend.iloc[0][12])  # Signal Category
                sheet1[SYSTEM_NUMBER] = dataappend.iloc[0][2]  # System number
                ##########################################################
                for i in range(0, len(dfreadtoui)):
                    sheet1.merge_cells(start_row=18 + i, start_column=3, end_row=18 + i,
                                       end_column=4)  # Merging C and D cols
                    sheet1.merge_cells(start_row=18 + i, start_column=5, end_row=18 + i,
                                       end_column=7)  # Merging E to G cols
                    sheet1.merge_cells(start_row=18 + i, start_column=8, end_row=18 + i,
                                       end_column=9)  # Merging H and I cols

                    self.set_border(sheet1, f'''B{18 + i}:I{18 + i}''')  # for excel sheet table cell borders

                    sheet1[f'''B{18 + i}'''] = i + 1
                    sheet1[f'''B{18 + i}'''].alignment = Alignment(
                        horizontal='center')  # Alignment of text center, left or right
                    sheet1[f'''C{18 + i}'''] = self.settestvalues.values[i]
                    sheet1[f'''C{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''E{18 + i}'''] = self.meastestvalues.values[i]
                    sheet1[f'''E{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''H{18 + i}'''] = self.errortestvalues.values[i]
                    sheet1[f'''H{18 + i}'''].alignment = Alignment(horizontal='center')

                    # merging cells for text at the end of sheet
                self.tableposition = 18 + len(dfreadtoui) + 2
                sheet1.merge_cells(start_row=self.tableposition, start_column=2, end_row=self.tableposition + 1,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 3, start_column=2, end_row=self.tableposition + 4,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 6, start_column=2, end_row=self.tableposition + 7,
                                   end_column=5)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=2, end_row=self.tableposition + 9,
                                   end_column=3)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=8, end_row=self.tableposition + 9,
                                   end_column=10)

                # excel writing
                sheet1[f'B{self.tableposition}'] = 'RMS Error :'
                sheet1[f'B{self.tableposition + 3}'] = 'Specification :'
                sheet1[f'B{self.tableposition + 6}'] = 'Status :'
                sheet1[f'B{self.tableposition + 8}'] = 'Tested By'
                sheet1[f'H{self.tableposition + 8}'] = 'Verified By'
                self.FontSizeStyle = [f'B{self.tableposition}', f'B{self.tableposition + 3}',
                                      f'B{self.tableposition + 6}',
                                      f'B{self.tableposition + 8}', f'H{self.tableposition + 8}']
                for i in range(0, len(self.FontSizeStyle)):
                    sheet1[self.FontSizeStyle[i]].font = Font(size=12, bold=True)  # font size and style
                # save the file
                outfile = excelfilename + '.xlsx'
                workbook.save(filename=f'{self.ReportPath}temp/' + outfile)
                print(dataappend)
            elif key_matched[i] == 'sensmeas':
                excelfilename = self.TableSelection[i]
                GetSensIndxTable = SensMeasIndxTableDB()
                GetSensIndxTable.GetSensRowRecord(select_record=self.TableSelection[i])
                dataappend = GetSensIndxTable.GetSensIndxRowRecord
                path = f'{self.ReportPath}Templates/SensMeasTestTemplate.xlsx'
                GetSensTestTable = SensMeasTestTableDB()
                GetSensTestTable.GetSensMeasTestTable(testtablename=self.TableSelection[i])
                dfreadtoui = GetSensTestTable.CurDbTableSens
                self.settestvalues = dfreadtoui["set_power"]
                self.meastestvalues = dfreadtoui["meas_sens"]
                self.errortestvalues = dfreadtoui["error"]
                dfRowFil = dataappend
                workbook = load_workbook(path)
                sheetonename = "sheet1"
                sheet1 = workbook.active
                sheet1.title = sheetonename
                sheet1[START_SENS] = dfRowFil.iloc[0][6]  # Start SENS
                sheet1[STOP_SENS] = (dfRowFil.iloc[0][7])  # Stop SENS
                sheet1[STEP_SENS] = (dfRowFil.iloc[0][8])  # Step SENS
                sheet1[SENS_POS_ANGLE] = (dfRowFil.iloc[0][11])  # Position Angle
                sheet1[SENS_FREQ] = (dfRowFil.iloc[0][13])  # PW
                sheet1[SENS_PRI] = (dfRowFil.iloc[0][14])  # PRI
                sheet1[SENS_SET_POWER] = (dfRowFil.iloc[0][10])  # Set Power
                datetrim = dataappend.iloc[0][0]
                dateformat = str(datetrim)[:10]
                timeformat = str(datetrim)[11:19]
                sheet1[TEST_SYSTEM] = (dataappend.iloc[0][3]).upper()  # system
                sheet1[TEST_MODE] = dataappend.iloc[0][4]  # mode
                sheet1[USER_NAME] = dataappend.iloc[0][1]  # name
                sheet1[TEST_DATE] = dateformat  # date
                sheet1[TEST_TIME] = timeformat  # time
                sheet1[SIGNAL_CATEGORY] = (dataappend.iloc[0][12])  # Signal Category
                sheet1[SYSTEM_NUMBER] = dataappend.iloc[0][2]  # System number
                for i in range(0, len(dfreadtoui)):
                    sheet1.merge_cells(start_row=18 + i, start_column=3, end_row=18 + i,
                                       end_column=4)  # Merging C and D cols
                    sheet1.merge_cells(start_row=18 + i, start_column=5, end_row=18 + i,
                                       end_column=7)  # Merging E to G cols
                    sheet1.merge_cells(start_row=18 + i, start_column=8, end_row=18 + i,
                                       end_column=9)  # Merging H and I cols

                    self.set_border(sheet1, f'''B{18 + i}:I{18 + i}''')  # for excel sheet table cell borders

                    sheet1[f'''B{18 + i}'''] = i + 1
                    sheet1[f'''B{18 + i}'''].alignment = Alignment(
                        horizontal='center')  # Alignment of text center, left or right
                    sheet1[f'''C{18 + i}'''] = self.settestvalues.values[i]
                    sheet1[f'''C{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''E{18 + i}'''] = self.meastestvalues.values[i]
                    sheet1[f'''E{18 + i}'''].alignment = Alignment(horizontal='center')
                    sheet1[f'''H{18 + i}'''] = self.errortestvalues.values[i]
                    sheet1[f'''H{18 + i}'''].alignment = Alignment(horizontal='center')

                    # merging cells for text at the end of sheet
                self.tableposition = 18 + len(dfreadtoui) + 2
                sheet1.merge_cells(start_row=self.tableposition, start_column=2, end_row=self.tableposition + 1,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 3, start_column=2, end_row=self.tableposition + 4,
                                   end_column=10)
                sheet1.merge_cells(start_row=self.tableposition + 6, start_column=2, end_row=self.tableposition + 7,
                                   end_column=5)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=2, end_row=self.tableposition + 9,
                                   end_column=3)
                sheet1.merge_cells(start_row=self.tableposition + 8, start_column=8, end_row=self.tableposition + 9,
                                   end_column=10)

                # excel writing
                sheet1[f'B{self.tableposition}'] = 'RMS Error :'
                sheet1[f'B{self.tableposition + 3}'] = 'Specification :'
                sheet1[f'B{self.tableposition + 6}'] = 'Status :'
                sheet1[f'B{self.tableposition + 8}'] = 'Tested By'
                sheet1[f'H{self.tableposition + 8}'] = 'Verified By'
                self.FontSizeStyle = [f'B{self.tableposition}', f'B{self.tableposition + 3}',
                                      f'B{self.tableposition + 6}',
                                      f'B{self.tableposition + 8}', f'H{self.tableposition + 8}']
                for i in range(0, len(self.FontSizeStyle)):
                    sheet1[self.FontSizeStyle[i]].font = Font(size=12, bold=True)  # font size and style
                # save the file
                outfile = excelfilename + '.xlsx'
                workbook.save(filename=f'{self.ReportPath}temp/' + outfile)
        self.PdfConvert(PdfsSelected=self.TableSelection)
    def PdfConvert(self, PdfsSelected=['esmfreqacctest_2024_04_30_15_50_45']):
        print("hello")
        self.TableSelection = PdfsSelected
        #########################################Converting To Pdf######################################################
        for i in range(0, len(self.TableSelection)):
            SheetFileName = str(self.TableSelection[i])
            excel = client.Dispatch("Excel.Application")
            # excel.Visible = False
            # Read Excel File
            sheets = excel.Workbooks.Open(f'{self.ReportPath}temp/{SheetFileName}.xlsx')
            excel.Visible = False
            work_sheets = sheets.Worksheets[0]
            # Convert into PDF File
            work_sheets.ExportAsFixedFormat(0,f'{self.ReportPath}temp/{SheetFileName}')  # no need for file format because it'll autoatically fixed
        print("pdf files generated Successfully and are ready for preview")
        #########################################File Merging###########################################################
        pdfs = []
        for i in range(0, len(self.TableSelection)):
            FilePathName = f'''C:/Users/jaswa/PycharmProjects/ATEC1/Reports/temp/{self.TableSelection[i]}.pdf'''
            pdfs.append(FilePathName)
        print(pdfs)
        merger = PdfMerger()
        for pdf in pdfs:
            merger.append(pdf)
        merger.write(f'{self.MainsrcPath}preview.pdf')
        merger.close()
        print("PDF merged successfully")
        map(lambda book: book.Close(True), excel.Workbooks)
        excel.Quit()
        self.CleanFiles()
        #########################################Temp File Clean Up#####################################################
    def CleanFiles(self):
        try:
            sleep(2)
            folder_path = f'{self.ReportPath}temp'
            shutil.rmtree(folder_path)
            print('Folder and its content removed')
        except:
            print('Folder not deleted')

        path = f'{self.ReportPath}temp'
        try:
            os.mkdir(path)
            print("Folder %s created!" % path)
        except FileExistsError:
            print("Folder %s already exists" % path)
        ################################################################################################################
    def set_border(self, worksheet, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in worksheet[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ####################################################################################################################
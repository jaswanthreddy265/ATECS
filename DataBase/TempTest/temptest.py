import datetime

import pandas as pd
from PyQt5.QtWidgets import QMessageBox
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Side, Border, Alignment, Font

from DataBase.INDXTBDB.AmplAccIndxTableDB import AmplAccIndxTableDB
from DataBase.INDXTBDB.DOAMeasIndxTableDB import DoaMeasIndxTableDB
from DataBase.INDXTBDB.FreqAccIndxTableDB import FreqAccIndxTableDB
from DataBase.INDXTBDB.PRIMeasIndxTableDB import PriMeasIndxTableDB
from DataBase.INDXTBDB.PWMeasIndxTableDB import PwMeasIndxTableDB
from DataBase.INDXTBDB.SensMeasAccIndxTableDB import SensMeasIndxTableDB
from DataBase.TESTTBDB.DoaMeasTestTableDB import DoaMeasTestTableDB
from DataBase.TESTTBDB.FreqAccTestTableDB import FreqAccTestTableDB
from DataBase.TESTTBDB.PriMeasTestTableDB import PriMeasTestTableDB
from DataBase.TESTTBDB.PwMeasTestTableDB import PwMeasTestTableDB


def SaveProcess(self):
    ####################################################
    GetFreqIndexTable = FreqAccIndxTableDB()
    GetFreqIndexTable.SelFreqDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                        datefilterto=self.dateEdit_To.text())
    freqdataappend = GetFreqIndexTable.FreqDateFil
    freqdataappend.insert(6, "testname", 'Frequency Test', True)
    freqdataappend = freqdataappend.astype(str)
    ########################
    GetPwIndexTable = PwMeasIndxTableDB()
    GetPwIndexTable.SelPwDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                    datefilterto=self.dateEdit_To.text())
    pwdataappend = GetPwIndexTable.PwDateFil
    pwdataappend.insert(6, "testname", 'Pulse Width Test', True)
    pwdataappend = pwdataappend.astype(str)
    ##################
    GetPriIndexTable = PriMeasIndxTableDB()
    GetPriIndexTable.SelPriDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                      datefilterto=self.dateEdit_To.text())
    pridataappend = GetPriIndexTable.PriDateFil
    pridataappend.insert(6, "testname", 'PRI Test', True)
    pridataappend = pridataappend.astype(str)
    ################
    GetAmplIndexTable = AmplAccIndxTableDB()
    GetAmplIndexTable.SelAmplDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                        datefilterto=self.dateEdit_To.text())
    ampldataappend = GetAmplIndexTable.AmplDateFil
    ampldataappend.insert(6, "testname", 'Amplitude Test', True)
    ampldataappend = ampldataappend.astype(str)
    #################
    GetDoaIndexTable = DoaMeasIndxTableDB()
    GetDoaIndexTable.SelDoaDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                      datefilterto=self.dateEdit_To.text())
    doadataappend = GetDoaIndexTable.DoaDateFil
    doadataappend.insert(6, "testname", 'DOA Test', True)
    doadataappend = doadataappend.astype(str)
    ####################
    GetSensMeasIndexTable = SensMeasIndxTableDB()
    GetSensMeasIndexTable.SelSensMeasDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                                datefilterto=self.dateEdit_To.text())
    sensmeasdataappend = GetSensMeasIndexTable.SensMeasDateFil
    sensmeasdataappend.insert(6, "testname", 'Sensitivity Test', True)
    sensmeasdataappend = sensmeasdataappend.astype(str)
    #######################
    dataappend = pd.concat(
        [freqdataappend, pwdataappend, pridataappend, ampldataappend, doadataappend, sensmeasdataappend],
        ignore_index=True)
    # print(dataappend)
    # dataappend.to_csv("freqaccindxtabledb.csv")
    selRows = self.tableWidget_Reports.selectedRanges()
    if (len(selRows) > 0):
        RowsSelected = []
        for i in range(len(selRows)):
            selRange = selRows[i]
            topRow = selRange.topRow()
            bottomRow = selRange.bottomRow()
            for row in range(topRow, bottomRow + 1):
                RowsSelected.append(row)
            RowsSelected.sort()
        # print(RowsSelected)

    TableSelection = []
    for i in range(len(RowsSelected)):
        rowofcol = (self.tableWidget_Reports.item(RowsSelected[i], 5)).text()
        TableSelection.append(rowofcol)
    print(TableSelection)

    for i in range(0, len(TableSelection)):
        ForTablename = TableSelection[i]
        if (self.CB_Test.currentText() == 'All'):
            GetFreqTestTable = FreqAccTestTableDB()
            GetFreqTestTable.GetFreqAccTestTable(testtablename=ForTablename)

            # print(GetFreqTestTable.CurDbTableFreq)
            dfreadtoui = GetFreqTestTable.CurDbTableFreq
            self.setfreqvalue = dfreadtoui["set_freq"]
            self.measfreqvalue = dfreadtoui["meas_freq"]
            self.errorvalue = dfreadtoui["error"]
            workbook = load_workbook(
                'C:/Users/jaswa/PycharmProjects/ATEC1/DataBase/Templates/FreqAccTestTemplate.xlsx')
            sheetonename = f'''{(TableSelection[i])}'''
            CurrentRow = self.tableWidget_Reports.currentRow()
            # print(sheetonename)
            sheet1 = workbook.active
            sheet1.title = sheetonename
            dfRowFil = dataappend.loc[dataappend['test_tab_ref'] == TableSelection[i]]
            # print(dfRowFil)
            datetrim = dfRowFil.iloc[0][0]
            dateformat = str(datetrim)[:10]
            timeformat = str(datetrim)[11:19]
            # print(timeformat)
            sheet1['F5'] = (dfRowFil.iloc[0][3]).upper()  # system
            sheet1['F6'] = dfRowFil.iloc[0][4]  # mode
            sheet1['E8'] = dfRowFil.iloc[0][1]  # name
            sheet1['E9'] = dateformat  # date
            sheet1['E12'] = dfRowFil.iloc[0][7]  # Start Freq
            sheet1['E13'] = (dfRowFil.iloc[0][8])  # Stop Freq
            sheet1['E14'] = (dfRowFil.iloc[0][9])  # Step Freq
            sheet1['E15'] = (dfRowFil.iloc[0][13])  # Signal Category
            sheet1['H8'] = dfRowFil.iloc[0][2]  # System number
            sheet1['H9'] = timeformat  # time
            sheet1['H12'] = (dfRowFil.iloc[0][12])  # Position Angle
            sheet1['H13'] = (dfRowFil.iloc[0][14])  # PW
            sheet1['H14'] = (dfRowFil.iloc[0][15])  # PRI
            sheet1['H15'] = (dfRowFil.iloc[0][11])  # Set Power

            for i in range(0, len(dfreadtoui)):
                sheet1.merge_cells(start_row=18 + i, start_column=3, end_row=18 + i,
                                   end_column=4)  # Merging C and D cols
                sheet1.merge_cells(start_row=18 + i, start_column=5, end_row=18 + i,
                                   end_column=7)  # Merging E to G cols
                sheet1.merge_cells(start_row=18 + i, start_column=8, end_row=18 + i,
                                   end_column=9)  # Merging H and I cols

                self.set_border(sheet1, f'''B{18 + i}:I{18 + i}''')  # for excel sheet table cell borders

                sheet1[f'''B{18 + i}'''] = i + 1
                sheet1[f'''B{18 + i}'''].alignment = Alignment(
                    horizontal='center')  # Alignment of text center, left or right
                sheet1[f'''C{18 + i}'''] = self.setfreqvalue.values[i]
                sheet1[f'''C{18 + i}'''].alignment = Alignment(horizontal='center')
                sheet1[f'''E{18 + i}'''] = self.measfreqvalue.values[i]
                sheet1[f'''E{18 + i}'''].alignment = Alignment(horizontal='center')
                sheet1[f'''H{18 + i}'''] = self.errorvalue.values[i]
                sheet1[f'''H{18 + i}'''].alignment = Alignment(horizontal='center')

            self.tableposition = 18 + len(dfreadtoui) + 2
            sheet1.merge_cells(start_row=self.tableposition, start_column=2, end_row=self.tableposition + 1,
                               end_column=10)
            sheet1.merge_cells(start_row=self.tableposition + 3, start_column=2, end_row=self.tableposition + 4,
                               end_column=10)
            sheet1.merge_cells(start_row=self.tableposition + 6, start_column=2, end_row=self.tableposition + 7,
                               end_column=5)
            sheet1.merge_cells(start_row=self.tableposition + 8, start_column=2, end_row=self.tableposition + 9,
                               end_column=3)
            sheet1.merge_cells(start_row=self.tableposition + 8, start_column=8, end_row=self.tableposition + 9,
                               end_column=10)

            sheet1[f'B{self.tableposition}'] = 'RMS Error :'
            sheet1[f'B{self.tableposition + 3}'] = 'Specification :'
            sheet1[f'B{self.tableposition + 6}'] = 'Status :'
            sheet1[f'B{self.tableposition + 8}'] = 'Tested By'
            sheet1[f'H{self.tableposition + 8}'] = 'Verified By'
            self.FontSizeStyle = [f'B{self.tableposition}', f'B{self.tableposition + 3}', f'B{self.tableposition + 6}',
                                  f'B{self.tableposition + 8}', f'H{self.tableposition + 8}']
            for i in range(0, len(self.FontSizeStyle)):
                sheet1[self.FontSizeStyle[i]].font = Font(size=12, bold=True)  # font size and style

            # save the file
            outfile = ForTablename + '.xlsx'
            workbook.save(filename=outfile)
            QMessageBox.information(self, "Self Test", "Reports Saved to " + outfile)
    ################################################################################################################

def set_border(self, worksheet, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in worksheet[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
import datetime
import random
import sys
from time import sleep

import pandas as pd
import pyqtgraph
from PyQt5 import QtCore
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog, QTableWidgetItem, QAbstractItemView, QMessageBox
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Side, Border

from DataBase.INDXTBDB.AmplAccIndxTableDB import AmplAccIndxTableDB
from DataBase.INDXTBDB.DOAMeasIndxTableDB import DoaMeasIndxTableDB
from DataBase.INDXTBDB.FreqAccIndxTableDB import FreqAccIndxTableDB
from DataBase.INDXTBDB.PRIMeasIndxTableDB import PriMeasIndxTableDB
from DataBase.INDXTBDB.PWMeasIndxTableDB import PwMeasIndxTableDB
from DataBase.INDXTBDB.SensMeasAccIndxTableDB import SensMeasIndxTableDB
from DataBase.TESTTBDB.DoaMeasTestTableDB import DoaMeasTestTableDB
from DataBase.TESTTBDB.FreqAccTestTableDB import FreqAccTestTableDB
from DataBase.TESTTBDB.PriMeasTestTableDB import PriMeasTestTableDB
from DataBase.TESTTBDB.PwMeasTestTableDB import PwMeasTestTableDB
from DataBase.MainSRC import Ui_ATEC_App
from DataBase.TESTTBDB.SensMeasTestTableDB import SensMeasTestTableDB


########################################################################################################################
class MainGUI(QMainWindow, Ui_ATEC_App):
    def __init__(self):
        super(MainGUI, self).__init__()
        pyqtgraph.setConfigOption('background', (255, 255, 255))
        self.setupUi(self)

        # Graphs Disabled for Network Setting
        self.frame_NET_ESMTracks.setDisabled(True)
        self.frame_NET_RFPS.setDisabled(True)
        self.frame_NET_ErrorGraphs.setDisabled(True)
        self.frame_NET_PolarPlot.setDisabled(True)


        # Reports Tab

        self.PB_Get.clicked.connect(self.ATECDataRetrieve)
        self.CB_Test.currentTextChanged.connect(self.SystemDisable)
        self.tableWidget_Reports.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableWidget_Reports.horizontalHeader().setStretchLastSection(True)
        #self.dateEdit_To.dateChanged.connect(self.ATECDataRetrieve)
        self.PlotGraph = self.Reports_ErrorGraph.addPlot()
#        self.PlotGraph = self.PlotGraph.setBackground("w")
        self.PB_Plot.clicked.connect(self.TableKeyValue)
        self.PB_Clear.clicked.connect(self.RemoveFilters)
        self.PB_Save.clicked.connect(self.SaveProcess)



        # Buttons
        self.PB_System.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.System_Parameters))
        self.PB_SubSystem.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.Subsystem_Parameters))
        self.PB_Complex.clicked.connect(lambda :self.stackedWidget.setCurrentWidget(self.ComplexEmiiter_Parameters))
        self.PB_Reports.clicked.connect(lambda :self.stackedWidget.setCurrentWidget(self.Reports_Parameters))
        self.PB_NetWork.clicked.connect(lambda :self.stackedWidget.setCurrentWidget(self.Network_Parameters))
        self.PB_View.clicked.connect(lambda :self.stackedWidget.setCurrentWidget(self.View_parameters))
        self.PB_Help.clicked.connect(lambda :self.stackedWidget.setCurrentWidget(self.Help_Parameters))

        self.PB_System.clicked.connect(lambda: self.stackedWidget_2.setCurrentWidget(self.System_Graphs))
        self.PB_SubSystem.clicked.connect(lambda: self.stackedWidget_2.setCurrentWidget(self.Subsystem_Graphs))
        self.PB_Complex.clicked.connect(lambda :self.stackedWidget_2.setCurrentWidget(self.ComplexEmitter_Graphs))
        self.PB_Reports.clicked.connect(lambda :self.stackedWidget_2.setCurrentWidget(self.Reports_Graphs))
        self.PB_NetWork.clicked.connect(lambda :self.stackedWidget_2.setCurrentWidget(self.Network_Graphs))
        self.PB_View.clicked.connect(lambda :self.stackedWidget_2.setCurrentWidget(self.View_Graphs))
        self.PB_Help.clicked.connect(lambda:self.stackedWidget_2.setCurrentWidget(self.Help_Docs))

        # File Browser
        self.PB_Report_PC.clicked.connect(self.file_open)

        # Graphs
        self.sys_error = self.SYS_ErrorGraph.addPlot(row=0, col=0)
        self.sys_polar = self.SYS_PolarPlot.addPlot(row=0, col=0)

        self.subsys_error = self.SUBSYS_ErrorGraph.addPlot(row=0, col=0)
        self.subsys_polar = self.SUBSYS_PolarPlot.addPlot(row=0, col=0)

        self.emitter_error = self.Emitter_ErrorGraph.addPlot(row=0, col=0)
        self.emitter_polar = self.Emitter_PolarPlot.addPlot(row=0, col=0)

        self.network_error = self.Network_ErrorGraph.addPlot(row=0, col=0)
        self.network_polar = self.Network_PolarPlot.addPlot(row=0, col=0)
        self.networ_rfps = self.Network_RFPSGraph.addPlot(row=0, col=0)


        self.view_error=self.View_ErrorGraph.addPlot(row=0, col=0)
        self.view_polar=self.View_PolarPlot.addPlot(row=0, col=0)
        #self.view_rfps=self.View_RFPSGraph.addPlot(row=0, col=0)

        self.Add.clicked.connect(self.Sys_Add)

        # Current date and time display to label
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.showTime)
        self.timer.start()

        self.graph_plot()

########################################################################################################################
    def showTime(self):
        self.time=QtCore.QDateTime.currentDateTime()
        self.timeDisplay=self.time.toString('yyyy-MM-dd hh:mm:ss dddd')
        self.label_date.setText(self.timeDisplay)

#######################################################################################################################
    def Sys_Add(self):
        if self.ESM.isChecked():
            if self.Injection.isChecked():

                if self.Frequency.isChecked():
                    self.listWidget_test.addItem("ESM Injection Frequency Test")
                    print("ESM Injection Frequency Test")

                elif self.Amplitude.isChecked():
                    self.listWidget_test.addItem("ESM Injection Amplitude Test")
                    print("ESM Injection Amplitude Test")

                elif self.Sensitivity.isChecked():
                    self.listWidget_test.addItem("ESM Injection Sensitivity Test")
                    print("ESM Injection Sensitivity Test")

                elif self.PW.isChecked():
                    self.listWidget_test.addItem("ESM Injection PW Test")
                    print("ESM Injection PW Test")

                elif self.PRI.isChecked():
                    self.listWidget_test.addItem("ESM Injection PRI Test")
                    print("ESM Injection PRI Test")

                elif self.DOA.isChecked():
                    self.listWidget_test.addItem("ESM Injection DOA Test")
                    print("ESM Injection DOA Test")
                else:
                    print("Select Test")

            elif self.Radiation.isChecked():

                if self.Frequency.isChecked():
                    self.listWidget_test.addItem("ESM Radiation Frequency Test")
                    print("ESM Radiation Frequency Test")

                elif self.Amplitude.isChecked():
                    self.listWidget_test.addItem("ESM Radiation Amplitude Test")
                    print("ESM Radiation Amplitude Test")

                elif self.Sensitivity.isChecked():
                    self.listWidget_test.addItem("ESM Radiation Sensitivity Test")
                    print("ESM Radiation Sensitivity Test")

                elif self.PW.isChecked():
                    self.listWidget_test.addItem("ESM Radiation PW Test")
                    print("ESM Radiation PW Test")

                elif self.PRI.isChecked():
                    self.listWidget_test.addItem("ESM Radiation PRI Test")
                    print("ESM Radiation PRI Test")

                elif self.DOA.isChecked():
                    self.listWidget_test.addItem("ESM Radiation DOA Test")
                    print("ESM Radiation DOA Test")

                else:
                    print("Select Test")

            else:
                print("Select any Mode")
        else:
            print("Select any System")

########################################################################################################################
    def file_open(self):
        filenames,_=QFileDialog.getOpenFileNames(self,"Select a file",'',"All Files (*)")
        if filenames:
            print(filenames)
            self.listWidget_Reports.addItems([str(filename) for filename in filenames ])
########################################################################################################################
########################################################################################################################
    def ATECDataRetrieve(self):
        self.tableWidget_Reports.setRowCount(0)
        if self.CB_Test.currentText() == 'All':
            self.Reports_ErrorGraph.hide()
            self.tableWidget_Reports.clear()
            self.tableWidget_Reports.setColumnCount(7)
            self.tableWidget_Reports.setHorizontalHeaderLabels(['date', 'username', 'system_id', 'system', 'mode', 'test_tab_ref', 'Test'])
            GetFreqIndexTable = FreqAccIndxTableDB()
            GetFreqIndexTable.SelFreqDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                                datefilterto=self.dateEdit_To.text())
            freqdataappend = GetFreqIndexTable.FreqDateFil
            freqdataappend.insert(6, "testname", 'Frequency Test', True)
            freqdataappend = freqdataappend.astype(str)
            #####
            GetPwIndexTable = PwMeasIndxTableDB()
            GetPwIndexTable.SelPwDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                            datefilterto=self.dateEdit_To.text())
            pwdataappend = GetPwIndexTable.PwDateFil
            pwdataappend.insert(6, "testname", 'Pulse Width Test', True)
            pwdataappend = pwdataappend.astype(str)
            ##################
            GetPriIndexTable = PriMeasIndxTableDB()
            GetPriIndexTable.SelPriDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                              datefilterto=self.dateEdit_To.text())
            pridataappend = GetPriIndexTable.PriDateFil
            pridataappend.insert(6, "testname", 'PRI Test', True)
            pridataappend = pridataappend.astype(str)
            ################
            GetAmplIndexTable = AmplAccIndxTableDB()
            GetAmplIndexTable.SelAmplDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                                datefilterto=self.dateEdit_To.text())
            ampldataappend = GetAmplIndexTable.AmplDateFil
            ampldataappend.insert(6, "testname", 'Amplitude Test', True)
            ampldataappend = ampldataappend.astype(str)
            #################
            GetDoaIndexTable = DoaMeasIndxTableDB()
            GetDoaIndexTable.SelDoaDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                              datefilterto=self.dateEdit_To.text())
            doadataappend = GetDoaIndexTable.DoaDateFil
            doadataappend.insert(6, "testname", 'DOA Test', True)
            doadataappend = doadataappend.astype(str)
            ####################
            GetSensMeasIndexTable = SensMeasIndxTableDB()
            GetSensMeasIndexTable.SelSensMeasDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                                        datefilterto=self.dateEdit_To.text())
            sensmeasdataappend = GetSensMeasIndexTable.SensMeasDateFil
            sensmeasdataappend.insert(6, "testname", 'Sensitivity Test', True)
            sensmeasdataappend = sensmeasdataappend.astype(str)
            #######################
            dataappend = pd.concat( [freqdataappend, pwdataappend, pridataappend, ampldataappend, doadataappend, sensmeasdataappend],
                ignore_index=True)
            print(dataappend)

        elif (self.CB_Test.currentText() == 'Frequency Test'):
            self.Reports_ErrorGraph.show()
            self.tableWidget_Reports.setColumnCount(6)
            self.tableWidget_Reports.clear()
            GetFreqIndexTable = FreqAccIndxTableDB()
            if (self.CB_System.currentText() == 'ESM') or (self.CB_System.currentText() == 'RFPS') or (
                    self.CB_System.currentText() == 'RWR') or (self.CB_System.currentText() == 'WARNER'):
                GetFreqIndexTable.SelFreqDateSysFilter(datefilterfrom=self.dateEdit_From.text(), datefilterto=self.dateEdit_To.text(), sysfilter=self.CB_System.currentText().lower())
                freqaccindxtabledata = GetFreqIndexTable.FreqDateSysFil
                # print(freqaccindxtabledata)
                dataappend = GetFreqIndexTable.FreqDateSysFil
            else:
                GetFreqIndexTable.SelFreqDateFilter(datefilterfrom=self.dateEdit_From.text(), datefilterto=self.dateEdit_To.text())
                freqaccindxtabledata = GetFreqIndexTable.FreqDateFil
                # print(freqaccindxtabledata)
                dataappend = GetFreqIndexTable.FreqDateFil
            self.PlotGraph.setTitle('<font size="15"><strong>Frequency Accuracy Test</strong></font>', color='blue')
            self.PlotGraph.setLabel( "left", '<span style="color: purple; font-size: 18px">Error </span>' )
            self.PlotGraph.setLabel( "bottom", '<span style="color: green; font-size: 18px">Set Frequency (MHz)</span>' )
        elif self.CB_Test.currentText() == 'PulseWidth Test':
            self.Reports_ErrorGraph.show()
            self.tableWidget_Reports.setColumnCount(6)
            self.tableWidget_Reports.clear()
            GetPwIndexTable = PwMeasIndxTableDB()
            if (self.CB_System.currentText() == 'ESM') or (self.CB_System.currentText() == 'RFPS') or (self.CB_System.currentText() == 'RWR') or (self.CB_System.currentText() == 'WARNER'):
                GetPwIndexTable.SelPwDateSysFilter(datefilterfrom=self.dateEdit_From.text(), datefilterto=self.dateEdit_To.text(), sysfilter=self.CB_System.currentText().lower())
                pwmeasindxtabledata = GetPwIndexTable.PwDateSysFil
                # print(pwmeasindxtabledata)
                dataappend = GetPwIndexTable.PwDateSysFil
            else:
                GetPwIndexTable.SelPwDateFilter(datefilterfrom=self.dateEdit_From.text(), datefilterto=self.dateEdit_To.text())
                pwmeasindxtabledata = GetPwIndexTable.PwDateFil
                # print(pwmeasindxtabledata)
                dataappend = GetPwIndexTable.PwDateFil
            self.PlotGraph.setTitle('<font size="15"><strong>Pulse Width Measurement Test</strong></font>', color='blue')
            self.PlotGraph.setLabel( "left", '<span style="color: purple; font-size: 18px">Error </span>' )
            self.PlotGraph.setLabel( "bottom", '<span style="color: green; font-size: 18px">Set Pulse Width (μs)</span>' )
        elif self.CB_Test.currentText() == 'PRI Test':
            self.Reports_ErrorGraph.show()
            self.tableWidget_Reports.setColumnCount(6)
            self.tableWidget_Reports.clear()
            GetPriIndexTable = PriMeasIndxTableDB()
            if (self.CB_System.currentText() == 'ESM') or (self.CB_System.currentText() == 'RFPS') or (self.CB_System.currentText() == 'RWR') or (self.CB_System.currentText() == 'WARNER'):
                GetPriIndexTable.SelPriDateSysFilter(datefilterfrom=self.dateEdit_From.text(), datefilterto=self.dateEdit_To.text(), sysfilter=self.CB_System.currentText().lower())
                primeasindxtabledata = GetPriIndexTable.PriDateSysFil
                # print(primeasindxtabledata)
                dataappend = GetPriIndexTable.PriDateSysFil
            else:
                GetPriIndexTable.SelPriDateFilter(datefilterfrom=self.dateEdit_From.text(), datefilterto=self.dateEdit_To.text())
                primeasindxtabledata = GetPriIndexTable.PriDateFil
                # print(primeasindxtabledata)
                dataappend = GetPriIndexTable.PriDateFil
            self.PlotGraph.setTitle('<font size="15"><strong>PRI Measurement Test</strong></font>', color='blue')
            self.PlotGraph.setLabel( "left", '<span style="color: purple; font-size: 18px">Error </span>' )
            self.PlotGraph.setLabel( "bottom", '<span style="color: green; font-size: 18px">Set Pri (μs)</span>' )
        elif self.CB_Test.currentText() == 'DOA Test':
            self.Reports_ErrorGraph.show()
            self.tableWidget_Reports.setColumnCount(6)
            self.tableWidget_Reports.clear()
            GetDoaIndexTable = DoaMeasIndxTableDB()
            if (self.CB_System.currentText() == 'ESM') or (self.CB_System.currentText() == 'RFPS') or ( self.CB_System.currentText() == 'RWR') or (self.CB_System.currentText() == 'WARNER'):
                GetDoaIndexTable.SelDoaDateSysFilter(datefilterfrom=self.dateEdit_From.text(), datefilterto=self.dateEdit_To.text(), sysfilter=self.CB_System.currentText().lower())
                doameasindxtabledata = GetDoaIndexTable.DoaDateSysFil
                # print(doameasindxtabledata)
                dataappend = GetDoaIndexTable.DoaDateSysFil
            else:
                GetDoaIndexTable.SelDoaDateFilter(datefilterfrom=self.dateEdit_From.text(), datefilterto=self.dateEdit_To.text())
                doameasindxtabledata = GetDoaIndexTable.DoaDateFil
                # print(doameasindxtabledata)
                dataappend = GetDoaIndexTable.DoaDateFil
            self.PlotGraph.setTitle('<font size="15"><strong>DOA Measurement Test</strong></font>', color='blue')
            self.PlotGraph.setLabel( "left", '<span style="color: purple; font-size: 18px">Error </span>' )
            self.PlotGraph.setLabel("bottom",'<span style="color: green; font-size: 18px">Set Angle (μs)</span>')

        elif self.CB_Test.currentText() == 'Amplitude Test':
            self.Reports_ErrorGraph.show()
            self.tableWidget_Reports.setColumnCount(6)
            self.tableWidget_Reports.clear()
            GetAmplIndexTable = AmplAccIndxTableDB()
            if (self.CB_System.currentText() == 'ESM') or (self.CB_System.currentText() == 'RFPS') or (
                    self.CB_System.currentText() == 'RWR') or (self.CB_System.currentText() == 'WARNER'):
                GetAmplIndexTable.SelAmplDateSysFilter(datefilterfrom=self.dateEdit_From.text(), datefilterto=self.dateEdit_To.text(), sysfilter=self.CB_System.currentText().lower())
                amplaccindxtabledata = GetAmplIndexTable.AmplDateSysFil
                # print(amplaccindxtabledata)
                dataappend = GetAmplIndexTable.AmplDateSysFil
            else:
                GetAmplIndexTable.SelAmplDateFilter(datefilterfrom=self.dateEdit_From.text(),datefilterto=self.dateEdit_To.text())
                amplaccindxtabledata = GetAmplIndexTable.AmplDateFil
                # print(amplaccindxtabledata)
                dataappend = GetAmplIndexTable.AmplDateFil
            self.PlotGraph.setTitle('<font size="15"><strong>Amplitude Accuracy Test</strong></font>', color='blue')
            self.PlotGraph.setLabel("left",'<span style="color: purple; font-size: 18px">Error </span>')
            self.PlotGraph.setLabel("bottom",'<span style="color: green; font-size: 18px">Set Amplitude (dBm)</span>')


        elif self.CB_Test.currentText() == 'Sensitivity Test':
            self.Reports_ErrorGraph.show()
            self.tableWidget_Reports.setColumnCount(6)
            self.tableWidget_Reports.clear()
            GetSensIndexTable = SensMeasIndxTableDB()
            if (self.CB_System.currentText() =='ESM') or (self.CB_System.currentText() =='RFPS') or (self.CB_System.currentText() =='RWR') or (self.CB_System.currentText() =='WARNER'):
                GetSensIndexTable.SelSensMeasDateSysFilter(datefilterfrom=self.dateEdit_From.text(),
                                                       datefilterto=self.dateEdit_To.text(),
                                                       sysfilter=self.CB_System.currentText().lower())
                sensmeasindxtabledata = GetSensIndexTable.SensMeasDateSysFil
                # print(sensmeasindxtabledata)
                dataappend = GetSensIndexTable.SensMeasDateSysFil
            else :
                GetSensIndexTable.SelSensMeasDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                                           datefilterto=self.dateEdit_To.text())
                sensmeasindxtabledata = GetSensIndexTable.SensMeasDateFil
                # print(sensmeasindxtabledata)
                dataappend = GetSensIndexTable.SensMeasDateFil
                self.PlotGraph.setTitle('<font size="15"><strong>Sensitivity Measurement Test</strong></font>',
                                        color='blue')
                self.PlotGraph.setLabel( "left", '<span style="color: purple; font-size: 18px">Error </span>' )
                self.PlotGraph.setLabel( "bottom", '<span style="color: green; font-size: 18px">Set Freq (MHz)</span>' )
        self.PlotGraph.clear()
        self.PlotGraph.setYRange(-100, 100)
        self.tableWidget_Reports.setHorizontalHeaderLabels(['date', 'username', 'system_id', 'system', 'mode', 'test_tab_ref'])
        dataappend = dataappend.astype(str)
        no_of_rows = dataappend.shape[0]
        no_of_cols = dataappend.shape[1]
        for row in range(no_of_rows):
            self.tableWidget_Reports.setRowCount(row + 1)
            for col in range(no_of_cols):
                self.tableWidget_Reports.setItem(row, col, QTableWidgetItem(dataappend.iloc[row][col]))

    ########################################################################################################################
    ########################################################################################################################
    def SystemDisable(self):
        if self.CB_Test.currentText() == "All":
            self.CB_System.setDisabled(True)
            self.PB_Plot.setDisabled(True)
        else:
            self.CB_System.setDisabled(False)
            self.PB_Plot.setDisabled(False)
    ########################################################################################################################

    def TableKeyValue(self):
        self.PlotGraph.clear()
        selRows = self.tableWidget_Reports.selectedRanges()
        if (len(selRows) > 0):
            # print('length of ranges',len(selRanges))
            # selRanges.sort()
            RowsSelected = []
            for i in range(len(selRows)):
                selRange = selRows[i]
                topRow = selRange.topRow()
                bottomRow = selRange.bottomRow()
                for row in range(topRow, bottomRow + 1):
                    RowsSelected.append(row)
                RowsSelected.sort()
            print(RowsSelected)

        TableSelection = []
        for i in range(len(RowsSelected)):
            rowofcol = (self.tableWidget_Reports.item(RowsSelected[i], 5)).text()
            TableSelection.append(rowofcol)
        print(TableSelection)

        for i in range(0, len(TableSelection)):
            ForTablename = TableSelection[i]
            if (self.CB_Test.currentText() == 'Frequency Test'):
                GetFreqTestTable = FreqAccTestTableDB()
                GetFreqTestTable.GetFreqAccTestTable(testtablename=ForTablename)
                print(GetFreqTestTable.CurDbTableFreq)
                dfreadtoui = GetFreqTestTable.CurDbTableFreq
                leftlabel = dfreadtoui["set_freq"].tolist()
                self.setfreqvalue = [float(i) for i in leftlabel]
                bottomlabel = dfreadtoui["error"].tolist()
                self.errorvalue = [float(i) for i in bottomlabel]
                self.PlotGraph.addLegend()
                self.PlotGraph.setYRange(-100, 100)
                self.rgb_full = (random.randint(1, 255), random.randint(1, 255), random.randint(1, 255))
                self.PlotGraph.plot(self.setfreqvalue, self.errorvalue,
                                    pen=pyqtgraph.mkPen(color=(self.rgb_full), width=2, style=QtCore.Qt.SolidLine),
                                    name=TableSelection[i])
            elif self.CB_Test.currentText() == 'PulseWidth Test':
                GetPwTestTable = PwMeasTestTableDB()
                GetPwTestTable.GetPwMeasTestTable(testtablename=ForTablename)
                print(GetPwTestTable.CurDbTablePw)
                dfreadtoui = GetPwTestTable.CurDbTablePw
                leftlabel = dfreadtoui["set_pw"].tolist()
                self.setpwvalue = [float(i) for i in leftlabel]
                bottomlabel = dfreadtoui["error"].tolist()
                self.errorvalue = [float(i) for i in bottomlabel]
                self.PlotGraph.addLegend()
                self.PlotGraph.setYRange(-100, 100)
                self.rgb_full = (random.randint(1, 255), random.randint(1, 255), random.randint(1, 255))
                self.PlotGraph.plot(self.setpwvalue, self.errorvalue,
                                    pen=pyqtgraph.mkPen(color=(self.rgb_full), width=2, style=QtCore.Qt.SolidLine),
                                    name=TableSelection[i])
            elif self.CB_Test.currentText() == 'DOA Test':
                GetDoaTestTable = DoaMeasTestTableDB()
                GetDoaTestTable.GetDoaMeasTestTable(testtablename=ForTablename)
                print(GetDoaTestTable.CurDbTableDoa)
                dfreadtoui = GetDoaTestTable.CurDbTableDoa
                leftlabel = dfreadtoui["set_angle"].tolist()
                self.setanglevalue = [float(i) for i in leftlabel]
                bottomlabel = dfreadtoui["error"].tolist()
                self.errorvalue = [float(i) for i in bottomlabel]
                self.PlotGraph.addLegend()
                self.PlotGraph.setYRange(-100, 100)
                self.rgb_full = (random.randint(1, 255), random.randint(1, 255), random.randint(1, 255))
                self.PlotGraph.plot(self.setanglevalue, self.errorvalue,
                                    pen=pyqtgraph.mkPen(color=(self.rgb_full), width=2, style=QtCore.Qt.SolidLine),
                                    name=TableSelection[i])
            elif self.CB_Test.currentText() == 'Sensitivity Test':
                GetSensTestTable = SensMeasTestTableDB()
                GetSensTestTable.GetSensMeasTestTable(testtablename=ForTablename)
                print(GetSensTestTable.CurDbTableSens)
                dfreadtoui = GetSensTestTable.CurDbTableSens
                leftlabel = dfreadtoui["set_freq"].tolist()
                self.setfreqvalue = [float(i) for i in leftlabel]
                bottomlabel = dfreadtoui["error"].tolist()
                self.errorvalue = [float(i) for i in bottomlabel]
                self.PlotGraph.addLegend()
                self.PlotGraph.setYRange(-100, 100)
                self.rgb_full = (random.randint(1, 255), random.randint(1, 255), random.randint(1, 255))
                self.PlotGraph.plot(self.setfreqvalue, self.errorvalue,
                                    pen=pyqtgraph.mkPen(color=(self.rgb_full), width=2, style=QtCore.Qt.SolidLine),
                                    name=TableSelection[i])

            elif self.CB_Test.currentText() == 'PRI Test':
                GetPriTestTable = PriMeasTestTableDB()
                GetPriTestTable.GetPriMeasTestTable(testtablename=ForTablename)
                print(GetPriTestTable.CurDbTablePri)
                dfreadtoui = GetPriTestTable.CurDbTablePri
                leftlabel = dfreadtoui["set_pri"].tolist()
                self.setprivalue = [float(i) for i in leftlabel]
                bottomlabel = dfreadtoui["error"].tolist()
                self.errorvalue = [float(i) for i in bottomlabel]
                self.PlotGraph.addLegend()
                self.PlotGraph.setYRange(-100, 100)
                self.rgb_full = (random.randint(1, 255), random.randint(1, 255), random.randint(1, 255))
                self.PlotGraph.plot(self.setprivalue, self.errorvalue,
                                    pen=pyqtgraph.mkPen(color=(self.rgb_full), width=2, style=QtCore.Qt.SolidLine),
                                    name=TableSelection[i])
    def RemoveFilters(self):
        self.tableWidget_Reports.clear()
        self.tableWidget_Reports.setRowCount(0)
        self.PlotGraph.clear()
    ####################################################################################################################
    def SaveProcess(self):
        ####################################################
        GetFreqIndexTable = FreqAccIndxTableDB()
        GetFreqIndexTable.SelFreqDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                            datefilterto=self.dateEdit_To.text())
        freqdataappend = GetFreqIndexTable.FreqDateFil
        freqdataappend.insert(6, "testname", 'Frequency Test', True)
        freqdataappend = freqdataappend.astype(str)
        ########################
        GetPwIndexTable = PwMeasIndxTableDB()
        GetPwIndexTable.SelPwDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                        datefilterto=self.dateEdit_To.text())
        pwdataappend = GetPwIndexTable.PwDateFil
        pwdataappend.insert(6, "testname", 'Pulse Width Test', True)
        pwdataappend = pwdataappend.astype(str)
        ##################
        GetPriIndexTable = PriMeasIndxTableDB()
        GetPriIndexTable.SelPriDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                          datefilterto=self.dateEdit_To.text())
        pridataappend = GetPriIndexTable.PriDateFil
        pridataappend.insert(6, "testname", 'PRI Test', True)
        pridataappend = pridataappend.astype(str)
        ################
        GetAmplIndexTable = AmplAccIndxTableDB()
        GetAmplIndexTable.SelAmplDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                            datefilterto=self.dateEdit_To.text())
        ampldataappend = GetAmplIndexTable.AmplDateFil
        ampldataappend.insert(6, "testname", 'Amplitude Test', True)
        ampldataappend = ampldataappend.astype(str)
        #################
        GetDoaIndexTable = DoaMeasIndxTableDB()
        GetDoaIndexTable.SelDoaDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                          datefilterto=self.dateEdit_To.text())
        doadataappend = GetDoaIndexTable.DoaDateFil
        doadataappend.insert(6, "testname", 'DOA Test', True)
        doadataappend = doadataappend.astype(str)
        ####################
        GetSensMeasIndexTable = SensMeasIndxTableDB()
        GetSensMeasIndexTable.SelSensMeasDateFilter(datefilterfrom=self.dateEdit_From.text(),
                                                    datefilterto=self.dateEdit_To.text())
        sensmeasdataappend = GetSensMeasIndexTable.SensMeasDateFil
        sensmeasdataappend.insert(6, "testname", 'Sensitivity Test', True)
        sensmeasdataappend = sensmeasdataappend.astype(str)
        #######################
        dataappend = pd.concat(
            [freqdataappend, pwdataappend, pridataappend, ampldataappend, doadataappend, sensmeasdataappend],
            ignore_index=True)
        # print(dataappend)
        #dataappend.to_csv("freqaccindxtabledb.csv")
        selRows = self.tableWidget_Reports.selectedRanges()
        if (len(selRows) > 0):
            RowsSelected = []
            for i in range(len(selRows)):
                selRange = selRows[i]
                topRow = selRange.topRow()
                bottomRow = selRange.bottomRow()
                for row in range(topRow, bottomRow + 1):
                    RowsSelected.append(row)
                RowsSelected.sort()
            # print(RowsSelected)

        TableSelection = []
        for i in range(len(RowsSelected)):
            rowofcol = (self.tableWidget_Reports.item(RowsSelected[i], 5)).text()
            TableSelection.append(rowofcol)
        print(TableSelection)

        for i in range(0, len(TableSelection)):
            ForTablename = TableSelection[i]
            if (self.CB_Test.currentText() == 'All'):
                GetFreqTestTable = FreqAccTestTableDB()
                GetFreqTestTable.GetFreqAccTestTable(testtablename=ForTablename)

                # print(GetFreqTestTable.CurDbTableFreq)
                dfreadtoui = GetFreqTestTable.CurDbTableFreq
                self.setfreqvalue = dfreadtoui["set_freq"]
                self.measfreqvalue = dfreadtoui["meas_freq"]
                self.errorvalue = dfreadtoui["error"]
                workbook = load_workbook(
                    'C:/Users/jaswa/PycharmProjects/ATEC1/DataBase/Templates/FreqAccTestTemplate.xlsx')
                sheetonename = f'''{(TableSelection[i])}'''
                CurrentRow = self.tableWidget_Reports.currentRow()
                # print(sheetonename)
                sheet1 = workbook.active
                sheet1.title = sheetonename
                dfRowFil = dataappend.loc[dataappend['test_tab_ref'] == TableSelection[i]]
                # print(dfRowFil)
                datetrim = dfRowFil.iloc[0][0]
                dateformat = str(datetrim)[:10]
                timeformat = str(datetrim)[11:19]
                # print(timeformat)
                sheet1['F4'] = (self.tableWidget_Reports.item(RowsSelected[i], 6)).text()
                sheet1['F5'] = (dfRowFil.iloc[0][3]).upper()
                sheet1['F6'] = dfRowFil.iloc[0][4]
                sheet1['E8'] = dfRowFil.iloc[0][1]
                sheet1['E9'] = dateformat
                sheet1['I8'] = dfRowFil.iloc[0][2]
                sheet1['I9'] = timeformat
                sheet1['E12'] = dfRowFil.iloc[0][7]
                sheet1['E13'] = (dfRowFil.iloc[0][9])
                sheet1['E14'] = (dfRowFil.iloc[0][12])
                sheet1['E15'] = (dfRowFil.iloc[0][14])
                sheet1['I12'] = (dfRowFil.iloc[0][8])
                sheet1['I13'] = (dfRowFil.iloc[0][11])
                sheet1['I14'] = (dfRowFil.iloc[0][13])
                sheet1['I15'] = (dfRowFil.iloc[0][15])

                sheet1['C17'] = 'Set Frequency(MHz)'
                sheet1['E17'] = 'Measure Frequency(MHz)'
                sheet1['I17'] = 'Error'

                for i in range(0, len(dfreadtoui)):
                    sheet1.merge_cells(start_row=18+i, start_column=3, end_row=18+i, end_column=4)
                    sheet1.merge_cells(start_row=18+i, start_column=5, end_row=18+i, end_column=8)
                    sheet1.merge_cells(start_row=18+i, start_column=9, end_row=18+i, end_column=10)
                    sheet1[f'''C{18+i}'''] = self.setfreqvalue.values[i]
                    sheet1[f'''E{18+i}'''] = self.measfreqvalue.values[i]
                    sheet1[f'''I{18+i}'''] = self.errorvalue.values[i]
                self.set_border(sheet1, f'''C18:I18''')  # for excel sheet table cell borders
                errorposition=18+len(dfreadtoui)+2
                sheet1.merge_cells(start_row=errorposition, start_column=f'B{errorposition}', end_row=errorposition, end_column=f'B{errorposition}')
                sheet1[f'B{errorposition}'] = 'RMS Error :'
                sheet1[f'B{errorposition+1}'] = 'Specification :'
                sheet1[f'B{errorposition+3}'] = 'Status :'
                sheet1[f'B{errorposition+5}'] = 'Tested By :'
                #sheet1[f'G{errorposition+5}'] = 'Verified By :'
                TestName = self.tableWidget_Reports.item(CurrentRow,6).text()
                # save the file
                outfile = ForTablename + '.xlsx'
                workbook.save(filename=outfile)
                QMessageBox.information(self, "Self Test", "Reports Saved to " + outfile)
        ################################################################################################################

    def set_border(self, worksheet, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in worksheet[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    def graph_plot(self):
        x1 = [2, 3, 4, 5]
        y1 = [10, 20, 15, 5]
        self.sys_error.plot(x1, y1, pen='r')
        x2 = [2, 3, 4, 5, 9, 15, 18, 20]
        y2 = [19, 22, 16, 67, 22, 34, 45, 100]
        self.sys_polar.plot(x2, y2, pen='y')

        x3 = [2, 3, 4, 5]
        y3 = [5, 25, 2, 45]
        self.subsys_error.plot(x3, y3, pen='m')
        x4 = [2, 3, 4, 5, 9, 15, 18, 20]
        y4 = [10, 20, 30, 15, 25, 35, 20, 5]
        self.subsys_polar.plot(x4, y4, pen='c')

        x5 = [2, 3, 4, 5]
        y5 = [5, 25, 2, 45]
        self.emitter_error.plot(x5, y5, pen='c')
        x6 = [2, 3, 4, 5, 9, 15, 18, 20]
        y6 = [10, 20, 30, 15, 25, 35, 20, 5]
        self.emitter_polar.plot(x6, y6, pen='g')

        """net_x=[2,3,5,7]
        net_error=[2,4,5,9]
        net_polar=[1,9,2,4]
        net_rfps=[1,9,2,6]
        self.network_error.plot(net_x,net_error,pen='g')
        self.network_polar.plot(net_x,net_polar,pen='b')
        self.networ_rfps.plot(net_x,net_rfps,pen='c')

        view_x = [2, 3, 5, 7,10]
        view_error = [10, 20, 30, 15, 25]
        view_polar = [20, 30, 15, 25, 35]
        view_rfps = [19, 22, 16, 67, 22]
        self.view_error.plot(view_x, view_error, pen='c')
        self.view_polar.plot(view_x, view_polar, pen='r')
        self.view_rfps.plot(view_x, view_rfps, pen='y')"""

########################################################################################################################

########################################################################################################################
if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = MainGUI()
    win.show()
    sys.exit(app.exec_())
